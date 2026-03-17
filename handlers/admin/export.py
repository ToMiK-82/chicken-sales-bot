"""
Команда /export — выгружает все заказы в XLSX с разделением по датам.
Доступна только админам.
✅ Исключает закрытые заказы (статусы issued, cancelled)
✅ Группирует заказы по дате поставки (поле date) в отдельные листы
✅ Подсвечивает статусы цветом: активный — зелёный, ожидание — жёлтый, отменён — красный, выдан — серый
✅ Отправляет сообщения ПОСЛЕ команды (не редактирует)
✅ Удаляет временное сообщение "Подготовка..." при необходимости
"""

from telegram import Update
from telegram.ext import ContextTypes, CommandHandler
from utils.admin_helpers import admin_required
from database.repository import db
from utils.formatting import format_phone
import logging
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import asyncio
from telegram.error import NetworkError, TimedOut
from collections import defaultdict

logger = logging.getLogger(__name__)

EXPORTS_DIR = "exports"
os.makedirs(EXPORTS_DIR, exist_ok=True)

HELP_TEXT = "📊 Выгрузить все заказы в Excel (XLSX) с группировкой по датам и цветовой маркировкой статусов"

# Определим цвета для статусов
STATUS_COLORS = {
    "active": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),      # зелёный
    "pending": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),     # жёлтый
    "cancelled": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),   # красный
    "issued": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),      # серый
}
STATUS_TEXT = {
    "active": "Активный",
    "cancelled": "Отменён",
    "issued": "Выдан",
    "pending": "Ожидает подтверждения"
}
# Статусы, которые считаются закрытыми (не включаем в выгрузку)
CLOSED_STATUSES = {"issued", "cancelled"}


@admin_required
async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выгружает заказы в XLSX, группируя по дате поставки"""
    effective_message = update.effective_message

    progress_msg = await effective_message.reply_text("⏳ Подготовка выгрузки...")

    filepath = None  # Чтобы было доступно в finally
    try:
        # Получаем все заказы
        rows = await db.execute_read("""
            SELECT id, breed, incubator, date, quantity, price, phone, status, created_at
            FROM orders ORDER BY created_at DESC
        """)

        if not rows:
            await effective_message.reply_text("❌ Нет заказов для выгрузки.")
            return

        # Группируем заказы по дате поставки
        grouped = defaultdict(list)
        for row in rows:
            status = row[7]
            if status in CLOSED_STATUSES:
                continue

            date_str = row[3] or ""
            delivery_date = date_str.split()[0] if date_str else "Без даты"
            grouped[delivery_date].append(row)

        if not grouped:
            await effective_message.reply_text("❌ Нет открытых заказов для выгрузки.")
            return

        total_count = sum(len(orders) for orders in grouped.values())
        logger.info(f"📊 Подготовлено {total_count} заказов для экспорта")

        # Создаём книгу Excel
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        currency_format = '# ##0 ₽'

        headers = ["Номер", "Порода", "Инкубатор", "Поставка", "Количество", "Цена, ₽", "Сумма, ₽", "Телефон", "Статус", "Создан"]

        for delivery_date, orders in grouped.items():
            sheet_title = delivery_date[:31] if delivery_date else "Без даты"
            ws = wb.create_sheet(title=sheet_title)

            ws.append(headers)
            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row in orders:
                order_id, breed, incubator, date, qty, price, phone, status, created_at = row

                try:
                    qty = int(qty)
                    price = int(float(price))
                    total = qty * price
                except (TypeError, ValueError):
                    qty = 0
                    price = 0
                    total = 0

                try:
                    delivery_date_fmt = datetime.strptime(date.split()[0], "%Y-%m-%d").strftime("%d.%m.%Y") if date else ""
                except:
                    delivery_date_fmt = date or ""

                try:
                    created_date = datetime.strptime(created_at.split()[0], "%Y-%m-%d").strftime("%d.%m.%Y") if created_at else ""
                except:
                    created_date = created_at or ""

                status_text = STATUS_TEXT.get(status, status.title())
                formatted_phone = format_phone(phone)

                ws.append([
                    order_id,
                    breed,
                    incubator or "Не указан",
                    delivery_date_fmt,
                    qty,
                    price,
                    total,
                    formatted_phone,
                    status_text,
                    created_date
                ])

                row_num = ws.max_row
                fill_color = STATUS_COLORS.get(status)
                if fill_color:
                    for col_num in range(1, 11):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.fill = fill_color

            column_widths = {
                'A': 8,   # Номер
                'B': 15,  # Порода
                'C': 18,  # Инкубатор
                'D': 12,  # Поставка
                'E': 10,  # Количество
                'F': 10,  # Цена
                'G': 12,  # Сумма
                'H': 18,  # Телефон
                'I': 12,  # Статус
                'J': 12   # Создан
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            for row_num in range(2, ws.max_row + 1):
                cell = ws[f'G{row_num}']
                cell.number_format = currency_format

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
                for cell in row:
                    cell.border = thin_border

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"orders_export_{timestamp}.xlsx"
        filepath = os.path.join(EXPORTS_DIR, filename)
        wb.save(filepath)
        logger.info(f"✅ XLSX-файл сохранён: {filepath}")

        # Отправка файла
        file_size = os.path.getsize(filepath)
        if file_size > 50 * 1024 * 1024:
            await effective_message.reply_text("❌ Файл слишком большой для отправки в Telegram.")
            # Не удаляем — пусть админ сам проверит
            logger.warning(f"⚠️ Файл слишком большой: {file_size} байт → {filepath}")
            return

        sent = False
        for attempt in range(3):
            try:
                with open(filepath, "rb") as f:
                    await effective_message.reply_document(
                        document=f,
                        filename=f"Заказы_{timestamp[:8]}.xlsx",
                        caption="📦 <b>Выгрузка заказов (открытые)</b>\n"
                                f"📊 Всего заказов: {total_count}\n"
                                f"🚫 Исключены: выданные и отменённые\n"
                                f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                        parse_mode="HTML"
                    )
                logger.info(f"📤 XLSX-экспорт отправлен после {attempt + 1} попыток")
                sent = True
                break
            except (NetworkError, TimedOut) as e:
                logger.warning(f"🔁 Попытка {attempt + 1} не удалась: {e}")
                if attempt < 2:
                    await asyncio.sleep(2 ** attempt)
            except Exception as e:
                logger.error(f"❌ Ошибка при отправке: {e}", exc_info=True)
                break

        if not sent:
            await effective_message.reply_text("❌ Не удалось отправить файл — ошибка сети.")
            logger.warning(f"⚠️ Экспорт не отправлен, файл сохранён: {filepath}")
        else:
            # Только при успехе — удаляем
            try:
                os.remove(filepath)
                logger.info(f"🗑️ Временный файл удалён: {filepath}")
            except Exception as e:
                logger.warning(f"⚠️ Не удалось удалить файл после отправки: {e}")

    except Exception as e:
        logger.error(f"❌ Ошибка при экспорте: {e}", exc_info=True)
        await effective_message.reply_text("❌ Не удалось создать выгрузку.")
        return

    finally:
        # Удаляем только сообщение "Подготовка..."
        try:
            await progress_msg.delete()
        except Exception as e:
            logger.debug(f"⚠️ Не удалось удалить сообщение 'Подготовка...': {e}")


def register_export_handler(application):
    application.add_handler(CommandHandler("export", export_command))
    logger.info("✅ Команда /export зарегистрирована")


def get_help_text() -> str:
    return HELP_TEXT
